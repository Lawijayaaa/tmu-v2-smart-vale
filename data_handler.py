#from pymodbus.client import ModbusSerialClient
from pymodbus.client import ModbusSerialClient
from toolboxTMU import parameter, sqlLibrary, find_tap, initParameter, dataParser, harmonicParser, convertBinList
from openpyxl import Workbook
import requests
import mysql.connector, time, datetime, math, openpyxl, sys, shutil, os
from requests.models import StreamConsumedError
from requests.exceptions import Timeout
import random

engineName = "Trafo B5T01"
teleURL = 'http://192.168.4.120:1444/api/transformer/sendNotificationToTelegramGroup'
progStat = True
debugMsg = False
infoMsg = True
dryType = False
gasType = False
transmitterModeMinus = False

exhibitStat = False
OLTCstat = False
pressureStat = True
tempStat = True

source = {
    16: 862,
    15: 817,
    14: 773,
    13: 728,
    12: 683,
    11: 639,
    10: 594,
    9: 548,
    8: 503,
    7: 458,
    6: 413,
    5: 367,
    4: 321,
    3: 276,
    2: 230,
    1: 184,
    0: 0
}

def main():
    if infoMsg == True: print("1D|Initialize Program") 
    dataLen = 56
    watchedData = 29
    CTratio = 1
    PTratio = 1
    eddyLosesGroup = 0.02
    designedKrated = 1
    loadCoef = 5
    cycleTime = 2 / 60
    
    client = ModbusSerialClient(method='rtu', port='/dev/ttyACM0', baudrate=9600)
    
    db = mysql.connector.connect(
        host = "localhost",
        user = "client",
        passwd = "raspi",
        database= "iot_trafo_client")
    cursor = db.cursor()

    #init logger rawdata
    ts = time.strftime("%Y%m%d")
    pathStr = r'/home/pi/tmu/tmu-app-client-deploy/assets/datalog/rawdata/datalogger-'
    #pathStr = r'/home/pi/tmu-v2-smart/assets/rawdata-test/datalogger-'
    pathDatLog = pathStr + ts + '.xlsx'
    sheetName = ["Harmonic_phR", "Harmonic_phS", "Harmonic_phT"]
    pathBkup = r'/home/pi/tmu-v2-smart/assets/rawdata-test/backup/datalogger-backup-'
    pathDatBkup = pathBkup + ts + '.xlsx'
     
    try:
        wb = openpyxl.load_workbook(pathDatLog)
        if infoMsg == True: print("1D|Open Existing Excel")
    except:
        #create new datalog
        workbook = Workbook()
        workbook.save(pathDatLog)
        #create datalog's header
        wb = openpyxl.load_workbook(pathDatLog)
        sheet = wb.active
        sheet.title = "Raw_data"
        name = (('timestamp', 
                    'V-un', 'V-vn', 'V-wn', 'V-uv', 'V-vw', 'V-uw',
                    'I-u', 'I-v', 'I-w', 'Itot', 'In',
                    'THDV-u', 'THDV-v', 'THDV-w', 'THDI-u', 'THDI-v', 'THDI-w',
                    'P-u', 'P-v', 'P-w', 'Ptot',
                    'Q-u', 'Q-v', 'Q-w', 'Q-tot',
                    'S-u', 'S-v', 'S-w', 'S-tot',
                    'PF-u', 'PF-v', 'PF-w', 'PFavg', 'Freq', 'kWh', 'kVARh',
                    'BusTemp-u', 'BusTemp-v', 'BusTemp-w', 'OilTemp',
                    'WTITemp-u', 'WTITemp-v', 'WTITemp-w',  'Press', 'Level',
                    'KRated-u', 'Derating-u', 'KRated-v', 'Derating-v', 'KRated-w', 'Derating-w',
                    'H2ppm', 'Moistppm', 'Vdiff-uv', 'Vdiff-vw', 'Vdiff-uw',
                    'trafoStatus', 'DIstat', 'DOstat', 'Alarm', 'Trip1', 'Trip2', 'Tap Position'),)
        for row in name:
            sheet.append(row)
        for member in sheetName:
            wb.create_sheet(member)
        for name in sheetName:
            sheetHarm = wb[name]
            rows = (('timestamp', 'V 1st', 'V 3rd' , 'V 5th' , 'V 7th' , 'V 9th' , 'V 11th' , 'V 13th' , 'V 15th' ,
                    'V 17th' , 'V 19th' , 'V 21st' , 'V 23rd' , 'V 25th' , 'V 27th' , 'V 29th' , 'V 31st',
                    'I 1st', 'I 3rd' , 'I 5th' , 'I 7th' , 'I 9th' , 'I 11th' , 'I 13th' , 'I 15th' ,
                    'I 17th' , 'I 19th' , 'I 21st' , 'I 23rd' , 'I 25th' , 'I 27th' , 'I 29th' , 'I 31st'),)
            for row in rows:
                sheetHarm.append(row)
        wb.save(pathDatLog)
        if debugMsg == True: print("1D|Create New Excel")
    
    inputData = [0]*dataLen
    currentStat = [0]*watchedData
    currentTrip = [0]*watchedData
    dataName = ['']*watchedData
    activeParam = [None]*watchedData
    activeFailure = [None]*watchedData
    kRated = [0, 0, 0]
    deRating = [0, 0, 0]
    hSquared = [0]*32
    timePassed = [0]*3
    deltaHi1 = [0]*3
    deltaHi2 = [0]*3
    deltaH1 = [0]*3
    deltaH2 = [0]*3
    lastLoadDefiner = [0]*3
    currentLoadDefiner = [0]*3
    raisingLoadBool = [True, True, True]
    loadFactor = [0]*3
    dataSet = [parameter("Name", 0, False, None, None, None, None, 3, 0)]
    for i in range(0, dataLen-1):
        dataSet.append(parameter("Name", 0, False, None, None, None, None, 3, 0))
    messageReason = ['Extreme Low',
                'Low', 
                'Back Normal', 
                'High', 
                'Extreme High']
    msgEvent = [None] * watchedData
    msgReminder = [None] * watchedData
    telePrevTime = excelPrevTime  = excelSavePrevTime = datetime.datetime.now()
    cursor.execute(sqlLibrary.sqlFailure)
    listFailure = cursor.fetchall()
    for i in range(0, len(listFailure)):
        if listFailure[i][2] == None:
            activeFailure[activeFailure.index(None)] = listFailure[i]
    
    if infoMsg == True: print("1D|Start Loop")
    while progStat:
        if debugMsg == True: print("1D|1 Fetch DB Data")
        start_time = time.time()
        cursor.execute(sqlLibrary.sqlTrafoSetting)
        trafoSetting = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlTrafoData)
        trafoData = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlTripSetting)
        tripSetting = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlDIscan)
        inputIO = cursor.fetchall()
        cursor.execute(sqlLibrary.sqlDOscan)
        outputIO = cursor.fetchall()
        cursor.execute(sqlLibrary.sqlConstantWTI, (str(trafoData[27]), ))
        constantWTI = cursor.fetchall()[0]
        cursor.execute(sqlLibrary.sqlTrafoStatus)
        prevStat = list(cursor.fetchall()[0][1:])
        cursor.execute(sqlLibrary.sqlTripStatus)
        prevTrip = list(cursor.fetchall()[0][1:])
        db.commit()
        CTratio = trafoData[26]

        if debugMsg == True: print("1D|3 Import Active Failure")
        if len(activeFailure):
            for i in range(0, len(activeFailure)):
                if activeFailure[i]:
                    activeParam[i] = activeFailure[i][4]
        
        if debugMsg == True: print("1D|3 Update status Relay")
        for i in range(0, 7):
            time.sleep(0.2)
            if outputIO[i][2] == 1:
                client.write_coil(i, True, slave = 1)
            elif outputIO[i][2] == 0:
                client.write_coil(i, False, slave = 1)
                
        if debugMsg == True: print("1D|4a Read Modbus Slave")
        time.sleep(0.2)
        getTemp = client.read_holding_registers(4, 3, slave = 3)
        time.sleep(0.2)
        getElect1 = client.read_holding_registers(0, 29, slave = 2)
        time.sleep(0.2)
        getElect2 = client.read_holding_registers(46, 5, slave = 2)
        time.sleep(0.2)
        getElect3 = client.read_holding_registers(800, 6, slave = 2)
        time.sleep(0.2)
        getHarmV = client.read_holding_registers(806, 90, slave = 2)
        time.sleep(0.2)
        getHarmI = client.read_holding_registers(896, 90, slave = 2)
        time.sleep(0.2)
        if gasType :
            getH2 = client.read_holding_registers(0, 1, slave = 4)
            getMoist = client.read_input_registers(0, 3, slave = 5)
            if debugMsg == True: print("1D|4b Parse Data")
            inputData = dataParser(exhibitStat, getTemp, getElect1, getElect2, getElect3, getH2, getMoist, dataLen, CTratio, PTratio)
        else :
            if debugMsg == True: print("1D|4b Parse Data")
            inputData = dataParser(exhibitStat, getTemp, getElect1, getElect2, getElect3, 0, 0, dataLen, CTratio, PTratio)
        if dryType :
            getTemp2 = client.read_holding_registers(0, 3, slave = 7)
            try :
                tempWTI = [0]*3
                for i in range(0, 3):
                    tempWTI[i] = (getTemp2.registers[i])/10
            except :
                pass
        
        if debugMsg == True: print("1D|5 Read Input IO")
        oilLevelAlarm = inputIO[4][2]
        oilLevelTrip = inputIO[5][2]
        analogIn1 = inputIO[6][2]
        analogIn2 = inputIO[7][2]

        if (oilLevelAlarm and oilLevelTrip) or oilLevelTrip:
            oilStat = 1
        elif oilLevelAlarm:
            oilStat = 2
        elif oilLevelAlarm == 0 and oilLevelTrip == 0:
            oilStat = 3
        inputData[44] = oilStat     #Oil Level

        if tempStat :
            if transmitterModeMinus :
                suhu = round(((analogIn1 * 0.009537) - 112.5), 3)
            else:
                suhu = round(((analogIn1 * 0.007630) - 50), 3)
            inputData[39] = max(0, suhu)
        else : 
            inputData[39] = 0
        
        if OLTCstat:
            tapPos = find_tap(round(analogIn2 * 0.06393945), source) + 1
            cursor.execute(sqlLibrary.sqlUpdateTapPos, (tapPos,))

        if pressureStat:
            inputData[43] = (analogIn2 - 6553) / 26214
        else:
            inputData[43] = 0
        
        #Exhibition only
        if exhibitStat:
            inputData[39] = (random.randint(3500, 5000))/100 #Oil Temp
            inputData[43] = (random.randint(10, 25))/100 #Pressure
            tapPos = random.randint(1, 15) #TapPos
            cursor.execute(sqlLibrary.sqlUpdateTapPos, (tapPos,))
        if dryType :
            if debugMsg == True: print("1D|6 Assign WTI")
            for i in range (0, 3) :
                inputData[i + 40] = tempWTI[i]
        else:
            if debugMsg == True: print("1D|6 Calculate WTI")
            for i in range(0, 3): loadFactor[i] = (inputData[i + 6])/trafoData[6]
            for i in range(0, 3):
                currentLoadDefiner[i] = inputData[i + 6]
                if currentLoadDefiner[i] - lastLoadDefiner[i] >= loadCoef:
                    timePassed[i] = 1
                    deltaHi1[i] = deltaH1[i]
                    deltaHi2[i] = deltaH2[i]
                    raisingLoadBool[i] = True
                    lastLoadDefiner[i] = currentLoadDefiner[i]
                elif lastLoadDefiner[i] - currentLoadDefiner[i] >= loadCoef:
                    timePassed[i] = 1
                    deltaHi1[i] = deltaH1[i]
                    deltaHi2[i] = deltaH2[i]
                    raisingLoadBool[i] = False
                    lastLoadDefiner[i] = currentLoadDefiner[i]
                else:
                    timePassed[i] = timePassed[i] + 1
                try:
                    if raisingLoadBool[i]:
                        deltaH1[i] = deltaHi1[i] + (((constantWTI[1] * trafoData[25] * trafoData[21]) * (math.pow(loadFactor[i], constantWTI[0])) - deltaHi1[i]) * (1 - math.exp((-1 * cycleTime * timePassed[i])/(constantWTI[2] * constantWTI[4]))))
                        deltaH2[i] = deltaHi2[i] + ((((constantWTI[1] - 1) * trafoData[25] * trafoData[21]) * (math.pow(loadFactor[i], constantWTI[0])) - deltaHi2[i]) * (1 - math.exp((-1 * cycleTime * timePassed[i] * constantWTI[2])/constantWTI[3])))
                        #print("rumus beban naik")
                    else:
                        deltaH1[i] = constantWTI[1] * trafoData[25] * trafoData[21] * math.pow(loadFactor[i], constantWTI[0]) + (deltaHi1[i] - (constantWTI[1] * trafoData[25] * trafoData[21] * math.pow(loadFactor[i], constantWTI[0]))) * (math.exp((-1 * cycleTime * timePassed[i])/(constantWTI[2] * constantWTI[4])))
                        deltaH2[i] = (constantWTI[1] - 1) * trafoData[25] * trafoData[21] * math.pow(loadFactor[i], constantWTI[0]) + (deltaHi2[i] - (constantWTI[1] - 1) * trafoData[25] * trafoData[21] * math.pow(loadFactor[i], constantWTI[0])) * math.exp(((-1 * cycleTime * timePassed[i] * constantWTI[4])/constantWTI[3]))
                        #print("rumus beban turun")
                except:
                    pass
                inputData[i + 40] = (round((inputData[39] + (deltaH1[i] - deltaH2[i])) * 100))/100
        
        if debugMsg == True: print("1D|7 Parse Harm, Update DB")
        inputHarmonicV = harmonicParser(getHarmV)
        inputHarmonicI = harmonicParser(getHarmI)
        cursor.execute(sqlLibrary.sqlUpdateVHarm1, inputHarmonicV[0])
        cursor.execute(sqlLibrary.sqlUpdateVHarm2, inputHarmonicV[1])
        cursor.execute(sqlLibrary.sqlUpdateVHarm3, inputHarmonicV[2])
        cursor.execute(sqlLibrary.sqlUpdateIHarm1, inputHarmonicI[0])
        cursor.execute(sqlLibrary.sqlUpdateIHarm2, inputHarmonicI[1])
        cursor.execute(sqlLibrary.sqlUpdateIHarm3, inputHarmonicI[2])
        if debugMsg == True: print("1D|8 Calculate KRated")
        kRatedlist = inputHarmonicI
        for i in range(0, 32):
            hSquared[i] = math.pow(((2*(i+1))-1), 2)
        for i in range(0, len(inputHarmonicI)):
            for j in range(0, len(inputHarmonicI[i])):
                kRatedlist[i][j] = math.pow((inputHarmonicI[i][j])/100, 2) * hSquared[j]
            kRated[i] = round(sum(kRatedlist[i]))
            deRating[i] = 100 * (math.pow((eddyLosesGroup + 1)/(kRated[i]*eddyLosesGroup + 1), 0.8) - math.pow((eddyLosesGroup + 1)/(designedKrated*eddyLosesGroup + 1), 0.8) + 1)
            if deRating[i] > 100:
                deRating[i] = 100
            inputData[i*2 + 45] = kRated[i]
            inputData[i*2 + 46] = (round(deRating[i] * 100))/100
        if debugMsg == True: print("1D|9 Input all data DB")
        dataResult = initParameter(dataSet, inputData, trafoSetting, trafoData, tripSetting, dataLen) 
        sendData = [datetime.datetime.now()] + inputData
        cursor.execute(sqlLibrary.sqlInsertData, sendData)
        db.commit()
        if debugMsg == True: print("1D|10 Check Failures Stat")
        maxStat = 0
        i =  0
        for data in dataResult:
            if data.isWatched:
                maxStat = data.trafoStat if data.trafoStat > maxStat else maxStat
                currentStat[i] = data.status
                currentTrip[i] = data.trafoStat
                dataName[i] = data.name
                #print(data.name)
                if data.status != prevStat[i]:
                    if data.status != 3:
                        if data.name in activeParam:
                            lastTimestamp = activeFailure[activeParam.index(data.name)][1]
                            duration = int((datetime.datetime.now() - lastTimestamp).total_seconds())
                            errorVal = [duration, activeFailure[activeParam.index(data.name)][0]]
                            cursor.execute(sqlLibrary.sqlResolveFailure, errorVal)
                            activeFailure[activeParam.index(data.name)] = None
                            activeParam[activeParam.index(data.name)] = None
                        errorVal = [datetime.datetime.now(), messageReason[data.status - 1], data.name, str(data.value)]
                        cursor.execute(sqlLibrary.sqlInsertFailure, errorVal)
                        cursor.execute(sqlLibrary.sqlLastFailure)
                        lastActive = cursor.fetchall()[0]
                        activeFailure[activeFailure.index(None)] = lastActive
                        loadProfile = str((round((data.value / trafoData[6]) * 10000))/100) + " Percent , Rated Current = " + str(trafoData[6])
                        msgEvent[i] = str(data.name + " " + messageReason[data.status - 1] + " , Value = " + (loadProfile if i == 3 or i == 4 or i == 5 else str(data.value)) + "\n" + "Time Occurence : " + str(datetime.datetime.now()))
                    elif data.status == 3:
                        lastTimestamp = activeFailure[activeParam.index(data.name)][1]
                        duration = int((datetime.datetime.now() - lastTimestamp).total_seconds())
                        errorVal = [duration, activeFailure[activeParam.index(data.name)][0]]
                        cursor.execute(sqlLibrary.sqlResolveFailure, errorVal)
                        activeFailure[activeParam.index(data.name)] = None
                        activeParam[activeParam.index(data.name)] = None
                        msgEvent[i] = None
                i = i + 1
        if debugMsg == True: print("1D|11 Check state changes")
        if prevStat != currentStat or prevTrip != currentTrip:
            #print("Send Telegram Lhooo")
            tele = list(filter(None, msgEvent))
            if tele:
                for message in tele:                
                    messages = engineName + " Says : " + "\n" + message
                    pload = {'message':messages}
                    try:
                        r = requests.post(teleURL, data = pload, timeout = 5, verify = False)
                    except Timeout:
                        pass
                    except Exception as Argument:
                        pass
            else:
                pass
            cursor.execute(sqlLibrary.sqlUpdateTransformerStatus, currentStat)
            cursor.execute(sqlLibrary.sqlUpdateTripStatus, currentTrip)
            cursor.execute(sqlLibrary.sqlUpdateTrafoStat, (maxStat,))
            db.commit()
        else:
            pass
        binList = convertBinList(inputIO, outputIO, currentTrip)
        if int((datetime.datetime.now() - telePrevTime).total_seconds()) > 3600:
            if debugMsg == True: print("1D|12 Routine remind Tele")
            #print("sekadar mengingatkan")
            for i in range(0, len(activeFailure)):
                if activeFailure[i]:
                    failureIndex = dataName.index(activeFailure[i][4])
                    msgReminder[failureIndex] = str(activeFailure[i][4] + " " + activeFailure[i][3] + " , Value = " + activeFailure[i][5] + "\n" + "Time Occurence : " + str(activeFailure[i][1]))                    
                    messages = engineName + " Says : " + "\n" + msgReminder
                    pload = {'message':messages}
                    try:
                        r = requests.post(teleURL, data = pload, timeout = 5, verify = False)
                    except Timeout:
                        pass
                    except Exception as Argument:
                        pass

            telePrevTime = datetime.datetime.now()
        #print(inputData)
        if int((datetime.datetime.now() - excelPrevTime).total_seconds()) > 3:
            if debugMsg == True: print("1D|12A Routine Add data to work stage excel")
            for i in range(0, 3):
                sendHarm = [datetime.datetime.now().strftime("%H:%M:%S")] + inputHarmonicV[i] + inputHarmonicI[i]
                sendHarm = ((tuple(sendHarm)),)
                sheetHarm = wb[sheetName[i]]
                for row in sendHarm:
                    sheetHarm.append(row)
            sendLog = [datetime.datetime.now().strftime("%H:%M:%S")] + inputData + [maxStat] + binList + [OLTCstat]
            sendLog = ((tuple(sendLog)),)
            sheet = wb["Raw_data"]
            for row in sendLog:
                sheet.append(row)
            excelPrevTime = datetime.datetime.now()
        if int((datetime.datetime.now() - excelSavePrevTime).total_seconds()) > 180:
            if debugMsg == True: print("1D|12B Routine Save Excel")
            if infoMsg == True: print("1D|Check Current Excel Size")
            if os.path.isfile(pathDatBkup) and os.path.getsize(pathDatBkup) >= os.path.getsize(pathDatLog):
                if infoMsg == True: print("1D|Excel Smaller than backup, replacing")
                shutil.copy2(pathDatBkup, pathDatLog)
            else:
                #create backup
                if infoMsg == True: print("1D|Backup Excel")
                shutil.copy2(pathDatLog, pathDatBkup)
            #print("save excel data here")
            try:
                if infoMsg == True: print("1D|Try to save Excel from work stage")
                wb.save(pathDatLog)
                time.sleep(0.5)
                if infoMsg == True: print("1D|Excel Size %s " % (os.path.getsize(pathDatLog)))
                if infoMsg == True: print("1D|Backup Size %s " % (os.path.getsize(pathDatBkup)))
                if (os.path.getsize(pathDatBkup) - os.path.getsize(pathDatLog)) < 3000:
                    if infoMsg == True: print("1D|Save Success")
                else:
                    raise Exception("backup larger than saved excel")
            except Exception as e:
                if infoMsg == True: print("1D|%s" % e)
                if infoMsg == True: print("1D|Save Failed, return to backup, restart system")
                shutil.copy2(pathDatBkup, pathDatLog)
                if infoMsg == True: print("1D|Restart")
            excelSavePrevTime = datetime.datetime.now()
                        
        cycleTime = (round(10000 * (time.time() - start_time)))/10000
        if debugMsg == True: print("1D|Cycle time %s" % cycleTime)
        print("1T|%s" % datetime.datetime.now())
        sys.stdout.flush()
        time.sleep(4)
        
if __name__ == "__main__":
    main()
